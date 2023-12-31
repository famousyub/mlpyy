from selenium import webdriver
from selenium.webdriver.common.by import By
from linkedin_scraper import Person,actions
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import time

import parameters




##Parameters that can be modified
startPage=1
nbPageToScrap=20
filename="recherche_candidat.xlsx"
workbook = load_workbook(filename = filename)
sheet= workbook.active
i=1
link=sheet.cell(i, 1).value
alreadyInData=[]
while(link):
    i+=1
    alreadyInData.append(link)
    link=sheet.cell(i, 1).value

##We intialize the variables containing the line number to write the next data
line=i

driver = webdriver.Chrome()
urlProfiles = ["https://www.linkedin.com/in/ayoub-smayen-619425222/","https://www.linkedin.com/in/mootez-meddeb-b35535107/"]
prospects = []
for url in urlProfiles:
    person = Person(url, driver=driver, scrape=False)
    try:
        element = WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.ID, "global-nav-search"))
            )
        person.scrape(close_on_complete=False)
        jobTitle = person.job_title.lower()
        experiences = person.experiences[0]
        print(jobTitle)
        print(person.name)
        print(experiences)
        if (jobTitle.find("web design") or jobTitle.find('designer web') or jobTitle.find('design web')):
            exp = re.search("from.*to Aujourd’hui for ([0-9]+) (jour|an|mois)", str(experiences))
            if ((exp.group(2) == "an" and int(exp.group(1)) < 2) or exp.group(2) == "jour" or exp.group(2) == "mois"):
                print("ok2")
                prospects.append([url, jobTitle, person.name, person.company])

                print("Writing in Excel")
                sheet.cell(line, 1).value = url
                sheet.cell(line, 2).value = jobTitle
                sheet.cell(line, 3).value = person.name
                sheet.cell(line, 4).value = person.company
                alreadyInData.append(url)
                workbook.save(filename=filename)
                line += 1
    except Exception as e:
        continue